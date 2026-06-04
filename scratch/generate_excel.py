#!/usr/bin/env python3
"""
Extract all Sugi Sushi menu data from menuData.ts and generate an organized Excel file.
"""
import json
import re
import os

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call(['pip3', 'install', 'openpyxl'])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# Read the TypeScript file
ts_path = os.path.join(os.path.dirname(__file__), '..', 'src', 'data', 'menuData.ts')
with open(ts_path, 'r', encoding='utf-8') as f:
    content = f.read()

# Extract the array portion
array_match = re.search(r'export const menuData:\s*Dish\[\]\s*=\s*\[', content)
if not array_match:
    raise ValueError("Could not find menuData array")

# Parse each object manually using regex
dishes = []
# Find all object blocks
object_pattern = re.compile(r'\{([^{}]*(?:\{[^{}]*\}[^{}]*)*)\}', re.DOTALL)

# Get the array content
array_start = array_match.end()
bracket_count = 1
array_end = array_start
for i in range(array_start, len(content)):
    if content[i] == '[':
        bracket_count += 1
    elif content[i] == ']':
        bracket_count -= 1
        if bracket_count == 0:
            array_end = i
            break

array_content = content[array_start:array_end]

# Parse each dish object
dish_blocks = []
brace_depth = 0
current_start = None
for i, ch in enumerate(array_content):
    if ch == '{':
        if brace_depth == 0:
            current_start = i
        brace_depth += 1
    elif ch == '}':
        brace_depth -= 1
        if brace_depth == 0 and current_start is not None:
            dish_blocks.append(array_content[current_start:i+1])
            current_start = None

def parse_value(val):
    val = val.strip()
    if val.startswith("'") and val.endswith("'"):
        return val[1:-1].replace("\\'", "'")
    elif val.startswith('"') and val.endswith('"'):
        return val[1:-1]
    elif val == 'true':
        return True
    elif val == 'false':
        return False
    elif val == '':
        return ''
    try:
        return int(val)
    except ValueError:
        try:
            return float(val)
        except ValueError:
            return val

def parse_array(text):
    """Parse a simple JS array like ['a', 'b', 'c']"""
    items = []
    text = text.strip()
    if text.startswith('[') and text.endswith(']'):
        text = text[1:-1].strip()
    if not text:
        return items
    # Split by comma, handling quoted strings
    current = ''
    in_quote = False
    quote_char = None
    for ch in text:
        if ch in ("'", '"') and not in_quote:
            in_quote = True
            quote_char = ch
        elif ch == quote_char and in_quote:
            in_quote = False
        elif ch == ',' and not in_quote:
            items.append(parse_value(current.strip()))
            current = ''
            continue
        current += ch
    if current.strip():
        items.append(parse_value(current.strip()))
    return items

def parse_portions(text):
    """Parse portions array"""
    portions = []
    # Find each object in the array
    brace_depth = 0
    current_start = None
    for i, ch in enumerate(text):
        if ch == '{':
            if brace_depth == 0:
                current_start = i
            brace_depth += 1
        elif ch == '}':
            brace_depth -= 1
            if brace_depth == 0 and current_start is not None:
                portion_text = text[current_start+1:i]
                portion = {}
                # Parse key-value pairs
                kvs = re.findall(r"(\w+)\s*:\s*(?:'([^']*)'|\"([^\"]*)\"|(\d+)|(\[[^\]]*\]))", portion_text)
                for kv in kvs:
                    key = kv[0]
                    value = kv[1] or kv[2] or kv[3] or kv[4]
                    if kv[3]:
                        value = int(value)
                    portion[key] = value
                portions.append(portion)
                current_start = None
    return portions

for block in dish_blocks:
    dish = {}
    inner = block[1:-1].strip()
    
    # Extract key-value pairs
    # Handle simple properties
    simple_props = re.findall(
        r"(\w+)\s*:\s*'((?:[^'\\]|\\.)*)'",
        inner
    )
    for key, val in simple_props:
        if key in ('name', 'nameAr', 'price', 'pieces'):
            continue  # Skip if inside portions
        dish[key] = val.replace("\\'", "'")
    
    # Better extraction - line by line
    dish = {}
    lines = inner.split('\n')
    for line in lines:
        line = line.strip().rstrip(',')
        if not line or line.startswith('//') or line.startswith('{') or line.startswith('}'):
            continue
        
        # Match key: 'value' or key: "value"
        m = re.match(r"(\w+)\s*:\s*'((?:[^'\\]|\\.)*)'$", line)
        if m:
            dish[m.group(1)] = m.group(2).replace("\\'", "'")
            continue
        
        m = re.match(r'(\w+)\s*:\s*"((?:[^"\\]|\\.)*)"$', line)
        if m:
            dish[m.group(1)] = m.group(2)
            continue
        
        # Match arrays: key: [...]
        m = re.match(r"(\w+)\s*:\s*(\[[^\]]*\])$", line)
        if m:
            dish[m.group(1)] = parse_array(m.group(2))
            continue
        
        # Match numbers
        m = re.match(r"(\w+)\s*:\s*(\d+)$", line)
        if m:
            dish[m.group(1)] = int(m.group(2))
            continue
    
    # Check for portions (multi-line array)
    portions_match = re.search(r'portions\s*:\s*\[(.*?)\]', block, re.DOTALL)
    if portions_match:
        dish['portions'] = parse_portions(portions_match.group(0))
    
    if 'id' in dish:
        dishes.append(dish)

print(f"Parsed {len(dishes)} dishes")

# Create Excel workbook
wb = openpyxl.Workbook()

# ===== STYLES =====
header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
title_font = Font(name='Calibri', bold=True, size=16, color='1F4E79')
subtitle_font = Font(name='Calibri', bold=True, size=11, color='2E75B6')
category_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
category_font = Font(name='Calibri', bold=True, size=11, color='1F4E79')
data_font = Font(name='Calibri', size=10)
price_font = Font(name='Calibri', size=10, bold=True, color='006600')
no_price_font = Font(name='Calibri', size=10, italic=True, color='999999')
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
alt_fill = PatternFill(start_color='F2F7FC', end_color='F2F7FC', fill_type='solid')
wrap_alignment = Alignment(wrap_text=True, vertical='center')
center_alignment = Alignment(horizontal='center', vertical='center')

def apply_header_style(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def apply_data_row_style(ws, row, max_col, is_alt=False):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if is_alt:
            cell.fill = alt_fill

# ===== SHEET 1: COMPLETE MENU =====
ws1 = wb.active
ws1.title = "Complete Menu"
ws1.sheet_properties.tabColor = "1F4E79"

# Headers
headers = ['#', 'ID', 'Category', 'Dish Name', 'Name (Arabic)', 'Description', 'Description (Arabic)', 
           'Price', 'Calories', 'Tags', 'Allergens', 'Portions Info']
for col, header in enumerate(headers, 1):
    ws1.cell(row=1, column=col, value=header)
apply_header_style(ws1, 1, len(headers))

# Data
row = 2
for idx, dish in enumerate(dishes, 1):
    ws1.cell(row=row, column=1, value=idx)
    ws1.cell(row=row, column=2, value=dish.get('id', ''))
    ws1.cell(row=row, column=3, value=dish.get('category', ''))
    ws1.cell(row=row, column=4, value=dish.get('name', ''))
    ws1.cell(row=row, column=5, value=dish.get('nameAr', ''))
    ws1.cell(row=row, column=6, value=dish.get('description', ''))
    ws1.cell(row=row, column=7, value=dish.get('descriptionAr', ''))
    
    price = dish.get('price', '')
    ws1.cell(row=row, column=8, value=price if price else 'N/A')
    if price:
        ws1.cell(row=row, column=8).font = price_font
    else:
        ws1.cell(row=row, column=8).font = no_price_font
    
    ws1.cell(row=row, column=9, value=dish.get('calories', ''))
    
    tags = dish.get('tags', [])
    ws1.cell(row=row, column=10, value=', '.join(tags) if isinstance(tags, list) else str(tags))
    
    allergens = dish.get('allergens', [])
    ws1.cell(row=row, column=11, value=', '.join(allergens) if isinstance(allergens, list) else str(allergens))
    
    # Portions
    portions = dish.get('portions', [])
    if portions:
        portions_str = '; '.join([f"{p.get('name', '')}: {p.get('price', '')} ({p.get('pieces', '')} pcs)" for p in portions])
        ws1.cell(row=row, column=12, value=portions_str)
    
    apply_data_row_style(ws1, row, len(headers), is_alt=(idx % 2 == 0))
    row += 1

# Set column widths
col_widths = [5, 18, 20, 30, 25, 50, 50, 10, 10, 35, 35, 40]
for i, width in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = width

# Freeze panes
ws1.freeze_panes = 'A2'

# Auto filter
ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row-1}"

# ===== SHEET 2: BY CATEGORY =====
ws2 = wb.create_sheet("By Category")
ws2.sheet_properties.tabColor = "2E75B6"

# Group dishes by category
from collections import OrderedDict
categories = OrderedDict()
for dish in dishes:
    cat = dish.get('category', 'Unknown')
    if cat not in categories:
        categories[cat] = []
    categories[cat].append(dish)

row = 1
for cat_name, cat_dishes in categories.items():
    # Category header
    ws2.cell(row=row, column=1, value=cat_name.upper())
    ws2.cell(row=row, column=1).font = category_font
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    for col in range(1, 9):
        ws2.cell(row=row, column=col).fill = category_fill
        ws2.cell(row=row, column=col).border = thin_border
    row += 1
    
    # Sub-headers
    sub_headers = ['#', 'Dish Name', 'Name (Arabic)', 'Description', 'Price', 'Calories', 'Tags', 'Allergens']
    for col, header in enumerate(sub_headers, 1):
        ws2.cell(row=row, column=col, value=header)
    apply_header_style(ws2, row, len(sub_headers))
    row += 1
    
    for idx, dish in enumerate(cat_dishes, 1):
        ws2.cell(row=row, column=1, value=idx)
        ws2.cell(row=row, column=2, value=dish.get('name', ''))
        ws2.cell(row=row, column=3, value=dish.get('nameAr', ''))
        ws2.cell(row=row, column=4, value=dish.get('description', ''))
        
        price = dish.get('price', '')
        ws2.cell(row=row, column=5, value=price if price else 'N/A')
        if price:
            ws2.cell(row=row, column=5).font = price_font
        else:
            ws2.cell(row=row, column=5).font = no_price_font
        
        ws2.cell(row=row, column=6, value=dish.get('calories', ''))
        
        tags = dish.get('tags', [])
        ws2.cell(row=row, column=7, value=', '.join(tags) if isinstance(tags, list) else str(tags))
        
        allergens = dish.get('allergens', [])
        ws2.cell(row=row, column=8, value=', '.join(allergens) if isinstance(allergens, list) else str(allergens))
        
        apply_data_row_style(ws2, row, len(sub_headers), is_alt=(idx % 2 == 0))
        row += 1
    
    row += 1  # Blank row between categories

# Column widths for sheet 2
col_widths2 = [5, 30, 25, 50, 10, 10, 35, 35]
for i, width in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = width

# ===== SHEET 3: PRICING SUMMARY =====
ws3 = wb.create_sheet("Pricing Summary")
ws3.sheet_properties.tabColor = "00B050"

headers3 = ['Category', 'Total Items', 'Items with Price', 'Items without Price', 'Lowest Price', 'Highest Price', 'Average Price']
for col, header in enumerate(headers3, 1):
    ws3.cell(row=1, column=col, value=header)
apply_header_style(ws3, 1, len(headers3))

row = 2
for cat_name, cat_dishes in categories.items():
    prices = []
    no_price_count = 0
    for dish in cat_dishes:
        price_str = dish.get('price', '')
        if price_str:
            # Extract numeric value
            nums = re.findall(r'[\d.]+', price_str)
            if nums:
                prices.append(float(nums[0]))
        else:
            no_price_count += 1
    
    ws3.cell(row=row, column=1, value=cat_name)
    ws3.cell(row=row, column=2, value=len(cat_dishes))
    ws3.cell(row=row, column=3, value=len(prices))
    ws3.cell(row=row, column=4, value=no_price_count)
    ws3.cell(row=row, column=5, value=f"{min(prices):.0f} SR" if prices else 'N/A')
    ws3.cell(row=row, column=6, value=f"{max(prices):.0f} SR" if prices else 'N/A')
    ws3.cell(row=row, column=7, value=f"{sum(prices)/len(prices):.1f} SR" if prices else 'N/A')
    
    apply_data_row_style(ws3, row, len(headers3), is_alt=(row % 2 == 0))
    row += 1

# Totals row
total_items = len(dishes)
all_prices = []
for dish in dishes:
    price_str = dish.get('price', '')
    if price_str:
        nums = re.findall(r'[\d.]+', price_str)
        if nums:
            all_prices.append(float(nums[0]))

ws3.cell(row=row, column=1, value='TOTAL')
ws3.cell(row=row, column=1).font = Font(name='Calibri', bold=True, size=11)
ws3.cell(row=row, column=2, value=total_items)
ws3.cell(row=row, column=3, value=len(all_prices))
ws3.cell(row=row, column=4, value=total_items - len(all_prices))
ws3.cell(row=row, column=5, value=f"{min(all_prices):.0f} SR" if all_prices else 'N/A')
ws3.cell(row=row, column=6, value=f"{max(all_prices):.0f} SR" if all_prices else 'N/A')
ws3.cell(row=row, column=7, value=f"{sum(all_prices)/len(all_prices):.1f} SR" if all_prices else 'N/A')

for col in range(1, len(headers3) + 1):
    cell = ws3.cell(row=row, column=col)
    cell.font = Font(name='Calibri', bold=True, size=11)
    cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    cell.border = thin_border

# Column widths for sheet 3
col_widths3 = [25, 12, 16, 18, 14, 14, 14]
for i, width in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = width

ws3.freeze_panes = 'A2'

# ===== SHEET 4: TAGS & ALLERGENS =====
ws4 = wb.create_sheet("Tags & Allergens")
ws4.sheet_properties.tabColor = "FFC000"

headers4 = ['Dish Name', 'Category', 'Price', 'Tags', 'Allergens']
for col, header in enumerate(headers4, 1):
    ws4.cell(row=1, column=col, value=header)
apply_header_style(ws4, 1, len(headers4))

row = 2
for idx, dish in enumerate(dishes, 1):
    ws4.cell(row=row, column=1, value=dish.get('name', ''))
    ws4.cell(row=row, column=2, value=dish.get('category', ''))
    ws4.cell(row=row, column=3, value=dish.get('price', '') or 'N/A')
    
    tags = dish.get('tags', [])
    ws4.cell(row=row, column=4, value=', '.join(tags) if isinstance(tags, list) else str(tags))
    
    allergens = dish.get('allergens', [])
    ws4.cell(row=row, column=5, value=', '.join(allergens) if isinstance(allergens, list) else str(allergens))
    
    apply_data_row_style(ws4, row, len(headers4), is_alt=(idx % 2 == 0))
    row += 1

col_widths4 = [30, 20, 10, 40, 40]
for i, width in enumerate(col_widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = width

ws4.freeze_panes = 'A2'
ws4.auto_filter.ref = f"A1:{get_column_letter(len(headers4))}{row-1}"

# ===== SHEET 5: ARABIC MENU =====
ws5 = wb.create_sheet("Arabic Menu (القائمة)")
ws5.sheet_properties.tabColor = "C00000"
ws5.sheet_view.rightToLeft = True

headers5 = ['#', 'الفئة', 'اسم الطبق', 'الوصف', 'السعر', 'السعرات']
for col, header in enumerate(headers5, 1):
    ws5.cell(row=1, column=col, value=header)
apply_header_style(ws5, 1, len(headers5))

row = 2
for idx, dish in enumerate(dishes, 1):
    ws5.cell(row=row, column=1, value=idx)
    ws5.cell(row=row, column=2, value=dish.get('category', ''))
    ws5.cell(row=row, column=3, value=dish.get('nameAr', '') or dish.get('name', ''))
    ws5.cell(row=row, column=4, value=dish.get('descriptionAr', '') or dish.get('description', ''))
    
    price = dish.get('price', '')
    ws5.cell(row=row, column=5, value=price if price else 'غير محدد')
    
    ws5.cell(row=row, column=6, value=dish.get('calories', ''))
    
    apply_data_row_style(ws5, row, len(headers5), is_alt=(idx % 2 == 0))
    row += 1

col_widths5 = [5, 20, 25, 50, 10, 10]
for i, width in enumerate(col_widths5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = width

ws5.freeze_panes = 'A2'

# ===== Save =====
output_path = os.path.join(os.path.dirname(__file__), '..', 'Sugi_Sushi_Complete_Menu.xlsx')
wb.save(output_path)
print(f"\n✅ Excel file saved to: {os.path.abspath(output_path)}")
print(f"\nSheets created:")
print(f"  1. Complete Menu - All {len(dishes)} dishes with full details")
print(f"  2. By Category - Dishes grouped by {len(categories)} categories")
print(f"  3. Pricing Summary - Price analysis per category")
print(f"  4. Tags & Allergens - Quick reference for dietary info")
print(f"  5. Arabic Menu (القائمة) - Arabic version of the menu")
